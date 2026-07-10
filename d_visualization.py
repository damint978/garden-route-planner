"""
角色D：可视化与前端官 - 静态图表生成脚本
严格对照赛题要求产出：
  ① 游客轨迹热力图（基于traffic_count）
  ② 图片热力图（基于photo_count）
  ③ 路线规划结果对比图（多条路线叠加）
   消融实验柱状图
  ⑤ 路线可解释性雷达图
数据来源：C部分输出的Excel文件
工具：folium（地图）、matplotlib（统计图）、plotly（雷达图）
"""

import pandas as pd
import numpy as np
import folium
from folium import plugins, Element
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import os

# ==================== 配置 ====================
EXCEL_PATH = r"f:\路线设计\c园林路径规划结果.xlsx"
OUTPUT_DIR = r"f:\路线设计\output"
os.makedirs(OUTPUT_DIR, exist_ok=True)

# 中文字体
plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei', 'Arial Unicode MS']
plt.rcParams['axes.unicode_minus'] = False

# 路线颜色
ROUTE_COLORS = {
    '最短路径': '#E63946',
    '最热路径': '#457B9D',
    '个性化路径-休闲放松型': '#2A9D8F',
}


# ==================== 数据加载 ====================
def load_data():
    import openpyxl
    wb = openpyxl.load_workbook(EXCEL_PATH)

    ws1 = wb['路径对比结果']
    paths = []
    headers1 = [c.value for c in ws1[1]]
    for row in ws1.iter_rows(min_row=2, values_only=True):
        paths.append(dict(zip(headers1, row)))

    ws2 = wb['消融实验结果']
    ablation = []
    headers2 = [c.value for c in ws2[1]]
    for row in ws2.iter_rows(min_row=2, values_only=True):
        ablation.append(dict(zip(headers2, row)))

    ws3 = wb['原始路段数据']
    roads = []
    headers3 = [c.value for c in ws3[1]]
    for row in ws3.iter_rows(min_row=2, values_only=True):
        roads.append(dict(zip(headers3, row)))

    df_roads = pd.DataFrame(roads)
    return paths, ablation, df_roads


def get_route_coords(road_ids_str, df_roads):
    """根据路段ID列表获取路线的经纬度坐标序列"""
    if not road_ids_str:
        return []
    road_ids = [rid.strip() for rid in str(road_ids_str).split(',')]
    coords = []
    for rid in road_ids:
        match = df_roads[df_roads['road_id'] == rid]
        if len(match) > 0:
            row = match.iloc[0]
            coords.append((row['start_lat'], row['start_lon']))
    if road_ids:
        last_rid = road_ids[-1].strip()
        match = df_roads[df_roads['road_id'] == last_rid]
        if len(match) > 0:
            row = match.iloc[0]
            coords.append((row['end_lat'], row['end_lon']))
    return coords


def get_route_scores(road_ids_str, df_roads):
    """计算路线各维度的平均评分"""
    if not road_ids_str:
        return {'文化': 0, '美观': 0, '拥堵': 0, '安静': 0}
    road_ids = [rid.strip() for rid in str(road_ids_str).split(',')]
    matched = df_roads[df_roads['road_id'].isin(road_ids)]
    if len(matched) == 0:
        return {'文化': 0, '美观': 0, '拥堵': 0, '安静': 0}
    return {
        '文化': round(matched['culture_score'].mean(), 1),
        '美观': round(matched['beauty_score'].mean(), 1),
        '拥堵': round(matched['congestion_score'].mean(), 1),
        '安静': round(matched['quiet_score'].mean(), 1),
    }


# ==================== ① 游客轨迹热力图 ====================
def draw_trajectory_heatmap(df_roads):
    """
    赛题要求：用户轨迹热力图
    基于traffic_count（游客流量）生成热力图，反映游客实际行走的热门区域
    """
    center_lat = df_roads['start_lat'].mean()
    center_lon = df_roads['start_lon'].mean()

    m = folium.Map(
        location=[center_lat, center_lon],
        zoom_start=16,
        tiles='CartoDB dark_matter',  # 深色底图，热力图更醒目
        attr='CartoDB'
    )

    # 热力数据：权重 = traffic_count（游客流量）
    max_traffic = df_roads['traffic_count'].max()
    heat_data = []
    for _, row in df_roads.iterrows():
        w = row['traffic_count'] / max_traffic
        heat_data.append([row['start_lat'], row['start_lon'], w])
        heat_data.append([row['end_lat'], row['end_lon'], w])

    plugins.HeatMap(
        heat_data,
        min_opacity=0.2,
        radius=18,
        blur=12,
        gradient={
            0.0: 'blue',
            0.2: 'cyan',
            0.4: 'lime',
            0.6: 'yellow',
            0.8: 'orange',
            1.0: 'red'
        }
    ).add_to(m)

    # 叠加路网（淡色）
    for _, row in df_roads.iterrows():
        folium.PolyLine(
            locations=[(row['start_lat'], row['start_lon']),
                       (row['end_lat'], row['end_lon'])],
            color='white',
            weight=0.8,
            opacity=0.12
        ).add_to(m)

    # 标题
    title_html = '''
    <div style="position: fixed; top: 10px; left: 50%; transform: translateX(-50%);
                z-index: 1000; background: rgba(0,0,0,0.75); color: white;
                padding: 10px 24px; border-radius: 8px; font-size: 16px;
                font-weight: bold; letter-spacing: 1px;">
        游客轨迹热力图（基于游客流量 traffic_count）
    </div>
    '''
    m.get_root().html.add_child(Element(title_html))

    output_path = os.path.join(OUTPUT_DIR, "01_游客轨迹热力图.html")
    m.save(output_path)
    print(f" ① 游客轨迹热力图 → {output_path}")
    return output_path


# ==================== ② 图片热力图 ====================
def draw_photo_heatmap(df_roads):
    """
    赛题要求：图片热力图
    基于photo_count（拍照数量）生成热力图，反映游客最喜爱的拍照打卡点
    """
    center_lat = df_roads['start_lat'].mean()
    center_lon = df_roads['start_lon'].mean()

    m = folium.Map(
        location=[center_lat, center_lon],
        zoom_start=16,
        tiles='CartoDB positron',  # 浅色底图
        attr='CartoDB'
    )

    max_photo = df_roads['photo_count'].max()
    heat_data = []
    for _, row in df_roads.iterrows():
        if row['photo_count'] > 0:
            w = row['photo_count'] / max_photo
            heat_data.append([row['start_lat'], row['start_lon'], w])
            heat_data.append([row['end_lat'], row['end_lon'], w])

    plugins.HeatMap(
        heat_data,
        min_opacity=0.15,
        radius=20,
        blur=14,
        gradient={
            0.0: '#a8d8ea',
            0.25: '#aa96da',
            0.5: '#fcbad3',
            0.75: '#ffffd2',
            1.0: '#ff6b6b'
        }
    ).add_to(m)

    # 标注TOP10拍照热点
    top10 = df_roads.nlargest(10, 'photo_count')
    for idx, (_, row) in enumerate(top10.iterrows()):
        folium.CircleMarker(
            location=[row['start_lat'], row['start_lon']],
            radius=10,
            color='#E63946',
            fill=True,
            fillColor='#E63946',
            fillOpacity=0.8,
            popup=f"TOP{idx+1} 拍照热点: {row['road_id']}<br>拍照数: {int(row['photo_count'])}",
        ).add_to(m)

    title_html = '''
    <div style="position: fixed; top: 10px; left: 50%; transform: translateX(-50%);
                z-index: 1000; background: rgba(0,0,0,0.75); color: white;
                padding: 10px 24px; border-radius: 8px; font-size: 16px;
                font-weight: bold;">
        图片热力图（基于拍照数量 photo_count）— 热门打卡点
    </div>
    '''
    m.get_root().html.add_child(Element(title_html))

    output_path = os.path.join(OUTPUT_DIR, "02_图片热力图.html")
    m.save(output_path)
    print(f" ② 图片热力图 → {output_path}")
    return output_path


# ==================== ③ 路线规划结果对比图 ====================
def draw_route_comparison(paths, df_roads):
    """
    赛题要求：路线规划结果对比图
    多条路线叠加在同一张地图上，清晰展示差异
    """
    all_coords = []
    for p in paths:
        coords = get_route_coords(p['路段ID列表'], df_roads)
        all_coords.extend(coords)

    center_lat = np.mean([c[0] for c in all_coords])
    center_lon = np.mean([c[1] for c in all_coords])

    m = folium.Map(
        location=[center_lat, center_lon],
        zoom_start=16,
        tiles='OpenStreetMap'
    )

    # 先画淡色路网底图
    for _, row in df_roads.iterrows():
        folium.PolyLine(
            locations=[(row['start_lat'], row['start_lon']),
                       (row['end_lat'], row['end_lon'])],
            color='#CCCCCC',
            weight=1,
            opacity=0.2
        ).add_to(m)

    # 绘制每条路线（带阴影效果）
    for p in paths:
        path_type = p['路径类型']
        coords = get_route_coords(p['路段ID列表'], df_roads)
        if not coords:
            continue

        color = ROUTE_COLORS.get(path_type, '#888888')

        # 阴影层（更粗、更淡）
        folium.PolyLine(
            locations=coords,
            color=color,
            weight=10,
            opacity=0.2,
        ).add_to(m)

        # 主线
        folium.PolyLine(
            locations=coords,
            color=color,
            weight=4,
            opacity=0.9,
            tooltip=f"{path_type} | 长度:{p['总长度(m)']:.0f}m | 评分:{p['总评分']:.1f}"
        ).add_to(m)

        # 起点
        folium.RegularPolygonMarker(
            location=coords[0],
            number_of_sides=3,
            radius=10,
            fill=True,
            fill_color=color,
            color='white',
            weight=2,
            popup=f"<b>{path_type}</b><br>起点"
        ).add_to(m)

        # 终点
        folium.RegularPolygonMarker(
            location=coords[-1],
            number_of_sides=4,
            radius=10,
            fill=True,
            fill_color=color,
            color='white',
            weight=2,
            popup=f"<b>{path_type}</b><br>终点"
        ).add_to(m)

    # 图例
    legend_html = '''
    <div style="position: fixed; bottom: 30px; left: 30px; z-index: 1000;
                background: white; padding: 14px 18px; border-radius: 10px;
                box-shadow: 0 3px 10px rgba(0,0,0,0.25); font-size: 13px;
                font-family: 'Microsoft YaHei', sans-serif;">
        <b style="font-size:14px;">路线图例</b><br><br>
        <span style="display:inline-block;width:30px;height:4px;background:#E63946;vertical-align:middle;margin-right:6px;"></span>
        最短路径（效率优先）<br>
        <span style="display:inline-block;width:30px;height:4px;background:#457B9D;vertical-align:middle;margin-right:6px;"></span>
        最热路径（热门打卡）<br>
        <span style="display:inline-block;width:30px;height:4px;background:#2A9D8F;vertical-align:middle;margin-right:6px;"></span>
        个性化路径（休闲放松）<br><br>
        <span style="color:#888;">&#9650;</span> 起点 &nbsp;
        <span style="color:#888;">&#9632;</span> 终点
    </div>
    '''
    m.get_root().html.add_child(Element(legend_html))

    title_html = '''
    <div style="position: fixed; top: 10px; left: 50%; transform: translateX(-50%);
                z-index: 1000; background: rgba(0,0,0,0.8); color: white;
                padding: 10px 24px; border-radius: 8px; font-size: 16px;
                font-weight: bold;">
        路线规划结果对比图（三条路线叠加）
    </div>
    '''
    m.get_root().html.add_child(Element(title_html))

    output_path = os.path.join(OUTPUT_DIR, "03_路线规划结果对比图.html")
    m.save(output_path)
    print(f" ③ 路线规划结果对比图 → {output_path}")
    return output_path


# ==================== ④ 消融实验柱状图 ====================
def draw_ablation_chart(ablation_data):
    """
    赛题要求：模型合理性 — 消融实验展示每一模块对模型效果的影响
    双图：评分变化 + 路径重合度
    """
    df_abl = pd.DataFrame(ablation_data)

    feature_names = {
        'culture_weight': '文化维度',
        'beauty_weight': '美观维度',
        'congestion_weight': '拥堵维度',
        'quiet_weight': '安静维度'
    }

    fig, axes = plt.subplots(2, 2, figsize=(16, 12))
    fig.suptitle('消融实验分析：各维度对路径规划的影响', fontsize=18, fontweight='bold', y=0.98)

    x_labels = [feature_names.get(f, f) for f in df_abl.iloc[:, 0]]
    x = np.arange(len(x_labels))
    width = 0.3

    # 左上图：基准 vs 消融后总评分
    ax1 = axes[0, 0]
    bars1 = ax1.bar(x - width / 2, df_abl.iloc[:, 1], width, label='基准总评分', color='#4ECDC4', alpha=0.9, edgecolor='white')
    bars2 = ax1.bar(x + width / 2, df_abl.iloc[:, 2], width, label='消融后总评分', color='#FF6B6B', alpha=0.9, edgecolor='white')
    ax1.set_title('(a) 总评分对比', fontsize=14, fontweight='bold')
    ax1.set_xticks(x)
    ax1.set_xticklabels(x_labels, fontsize=11)
    ax1.set_ylabel('总评分', fontsize=12)
    ax1.legend(fontsize=10)
    ax1.grid(axis='y', alpha=0.3)
    for bar in bars1:
        ax1.annotate(f'{bar.get_height():.0f}', xy=(bar.get_x() + bar.get_width()/2, bar.get_height()),
                     xytext=(0, 3), textcoords="offset points", ha='center', fontsize=9)
    for bar in bars2:
        ax1.annotate(f'{bar.get_height():.0f}', xy=(bar.get_x() + bar.get_width()/2, bar.get_height()),
                     xytext=(0, 3), textcoords="offset points", ha='center', fontsize=9)

    # 右上图：评分变化率
    ax2 = axes[0, 1]
    change_rates = df_abl.iloc[:, 3]
    colors_rate = ['#FF6B6B' if v > 0 else '#4ECDC4' for v in change_rates]
    bars3 = ax2.bar(x, change_rates, width=0.5, color=colors_rate, alpha=0.85, edgecolor='white')
    ax2.set_title('(b) 评分变化率 (%)', fontsize=14, fontweight='bold')
    ax2.set_xticks(x)
    ax2.set_xticklabels(x_labels, fontsize=11)
    ax2.set_ylabel('变化率 (%)', fontsize=12)
    ax2.axhline(y=0, color='black', linewidth=0.8)
    ax2.grid(axis='y', alpha=0.3)
    for bar in bars3:
        h = bar.get_height()
        ax2.annotate(f'{h:.1f}%', xy=(bar.get_x() + bar.get_width()/2, h),
                     xytext=(0, 3 if h >= 0 else -12), textcoords="offset points",
                     ha='center', fontsize=10, fontweight='bold')

    # 左下图：路径重合度
    ax3 = axes[1, 0]
    overlap = df_abl.iloc[:, 4]
    colors_ov = ['#FF6B6B' if v < 100 else '#4ECDC4' for v in overlap]
    bars4 = ax3.bar(x, overlap, width=0.5, color=colors_ov, alpha=0.85, edgecolor='white')
    ax3.set_title('(c) 路径重合度 (%)', fontsize=14, fontweight='bold')
    ax3.set_xticks(x)
    ax3.set_xticklabels(x_labels, fontsize=11)
    ax3.set_ylabel('重合度 (%)', fontsize=12)
    ax3.set_ylim(0, 110)
    ax3.grid(axis='y', alpha=0.3)
    for bar in bars4:
        ax3.annotate(f'{bar.get_height():.0f}%', xy=(bar.get_x() + bar.get_width()/2, bar.get_height()),
                     xytext=(0, 3), textcoords="offset points", ha='center', fontsize=11, fontweight='bold')

    # 右下图：路段数对比
    ax4 = axes[1, 1]
    bars5 = ax4.bar(x - width / 2, df_abl.iloc[:, 5], width, label='基准路段数', color='#4ECDC4', alpha=0.9)
    bars6 = ax4.bar(x + width / 2, df_abl.iloc[:, 6], width, label='消融后路段数', color='#FF6B6B', alpha=0.9)
    ax4.set_title('(d) 路段数对比', fontsize=14, fontweight='bold')
    ax4.set_xticks(x)
    ax4.set_xticklabels(x_labels, fontsize=11)
    ax4.set_ylabel('路段数', fontsize=12)
    ax4.legend(fontsize=10)
    ax4.grid(axis='y', alpha=0.3)

    plt.tight_layout(rect=[0, 0, 1, 0.95])
    output_path = os.path.join(OUTPUT_DIR, "04_消融实验柱状图.png")
    fig.savefig(output_path, dpi=200, bbox_inches='tight', facecolor='white')
    plt.close(fig)
    print(f" ④ 消融实验柱状图 → {output_path}")
    return output_path


# ==================== ⑤ 路线可解释性雷达图 ====================
def draw_radar_chart(paths, df_roads):
    """
    赛题要求：模型可解释性 — 以可视化形式解释规划出的路线如何满足目标函数
    用雷达图展示三条路线在四维度上的表现差异
    """
    categories = ['文化', '美观', '拥堵', '安静']
    N = len(categories)
    angles = [n / float(N) * 2 * np.pi for n in range(N)]
    angles += angles[:1]

    fig, ax = plt.subplots(figsize=(8, 8), subplot_kw=dict(polar=True))

    for p in paths:
        scores = get_route_scores(p['路段ID列表'], df_roads)
        values = [scores[c] for c in categories]
        values += values[:1]
        color = ROUTE_COLORS.get(p['路径类型'], '#888888')
        ax.plot(angles, values, 'o-', linewidth=2.5, label=p['路径类型'], color=color, markersize=8)
        ax.fill(angles, values, alpha=0.1, color=color)

    ax.set_xticks(angles[:-1])
    ax.set_xticklabels(categories, fontsize=14)
    ax.set_ylim(0, 100)
    ax.set_yticks([20, 40, 60, 80, 100])
    ax.set_yticklabels(['20', '40', '60', '80', '100'], fontsize=10, color='gray')
    ax.set_title('路线可解释性：四维度评分雷达图', fontsize=16, fontweight='bold', pad=20)
    ax.legend(loc='upper right', bbox_to_anchor=(1.3, 1.1), fontsize=12)
    ax.grid(True, alpha=0.3)

    plt.tight_layout()
    output_path = os.path.join(OUTPUT_DIR, "05_路线可解释性雷达图.png")
    fig.savefig(output_path, dpi=200, bbox_inches='tight', facecolor='white')
    plt.close(fig)
    print(f" ⑤ 路线可解释性雷达图 → {output_path}")
    return output_path


# ==================== ⑥ 路线详细指标对比柱状图 ====================
def draw_route_metrics_comparison(paths, df_roads):
    """三条路线在长度、评分、路段数等指标上的对比"""
    fig, axes = plt.subplots(1, 3, figsize=(18, 6))
    fig.suptitle('三条路线规划方案综合对比', fontsize=16, fontweight='bold')

    labels = [p['路径类型'].replace('个性化路径-', '') for p in paths]
    colors = [ROUTE_COLORS.get(p['路径类型'], '#888') for p in paths]

    # 总长度对比
    lengths = [p['总长度(m)'] for p in paths]
    ax1 = axes[0]
    bars1 = ax1.bar(labels, lengths, color=colors, alpha=0.85, edgecolor='white', width=0.5)
    ax1.set_title('总长度对比 (m)', fontsize=13, fontweight='bold')
    ax1.set_ylabel('长度 (m)', fontsize=12)
    for bar in bars1:
        ax1.annotate(f'{bar.get_height():.0f}m', xy=(bar.get_x() + bar.get_width()/2, bar.get_height()),
                     xytext=(0, 3), textcoords="offset points", ha='center', fontsize=11, fontweight='bold')
    ax1.grid(axis='y', alpha=0.3)

    # 总评分对比
    scores = [p['总评分'] for p in paths]
    ax2 = axes[1]
    bars2 = ax2.bar(labels, scores, color=colors, alpha=0.85, edgecolor='white', width=0.5)
    ax2.set_title('总评分对比', fontsize=13, fontweight='bold')
    ax2.set_ylabel('评分', fontsize=12)
    for bar in bars2:
        ax2.annotate(f'{bar.get_height():.1f}', xy=(bar.get_x() + bar.get_width()/2, bar.get_height()),
                     xytext=(0, 3), textcoords="offset points", ha='center', fontsize=11, fontweight='bold')
    ax2.grid(axis='y', alpha=0.3)

    # 路段数对比
    segs = [p['总路段数'] for p in paths]
    ax3 = axes[2]
    bars3 = ax3.bar(labels, segs, color=colors, alpha=0.85, edgecolor='white', width=0.5)
    ax3.set_title('路段数对比', fontsize=13, fontweight='bold')
    ax3.set_ylabel('路段数', fontsize=12)
    for bar in bars3:
        ax3.annotate(f'{int(bar.get_height())}', xy=(bar.get_x() + bar.get_width()/2, bar.get_height()),
                     xytext=(0, 3), textcoords="offset points", ha='center', fontsize=11, fontweight='bold')
    ax3.grid(axis='y', alpha=0.3)

    plt.tight_layout()
    output_path = os.path.join(OUTPUT_DIR, "06_路线指标对比柱状图.png")
    fig.savefig(output_path, dpi=200, bbox_inches='tight', facecolor='white')
    plt.close(fig)
    print(f" ⑥ 路线指标对比柱状图 → {output_path}")
    return output_path


# ==================== 主函数 ====================
def main():
    print("=" * 60)
    print("  角色D：可视化与前端官 — 静态图表生成")
    print("  严格对照赛题：轨迹热力图、图片热力图、路线对比图、")
    print("              消融实验、可解释性可视化")
    print("=" * 60)

    paths, ablation, df_roads = load_data()
    print(f"\n数据加载完成：{len(paths)} 条路径, {len(ablation)} 条消融记录, {len(df_roads)} 条路段\n")

    draw_trajectory_heatmap(df_roads)
    draw_photo_heatmap(df_roads)
    draw_route_comparison(paths, df_roads)
    draw_ablation_chart(ablation)
    draw_radar_chart(paths, df_roads)
    draw_route_metrics_comparison(paths, df_roads)

    print(f"\n{'=' * 60}")
    print(f"  全部 6 张图表生成完毕！")
    print(f"  输出目录：{OUTPUT_DIR}")
    print(f"{'=' * 60}")


if __name__ == "__main__":
    main()
